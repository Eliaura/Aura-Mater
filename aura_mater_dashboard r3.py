"""
AURA Energy — MATER Dashboard
Lee los Excel de Cammesa desde data/. Para actualizar cada trimestre:
reemplazá los dos archivos en la carpeta data/ del repo de GitHub.
"""
import re
import streamlit as st
import pandas as pd
import plotly.express as px

st.set_page_config(
    page_title="AURA Energy — MATER",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
[data-testid="stSidebar"] { background: #0F1923; }
[data-testid="stSidebar"] * { color: #E8F0FE !important; }
.metric-card {
    background: white; border-radius: 10px; padding: 14px 18px;
    border-left: 4px solid #2E75B6; box-shadow: 0 1px 4px rgba(0,0,0,0.08);
}
.metric-val { font-size: 1.9rem; font-weight: 600; color: #1F3864; line-height: 1.1; }
.metric-lbl { font-size: 0.72rem; color: #666; text-transform: uppercase; letter-spacing:.06em; }
</style>
""", unsafe_allow_html=True)

# ── RUTAS A LOS EXCEL ─────────────────────────────────────────────────────────
REFA_PATH  = "data/MATER_Referencial_A_1T2026_18.xlsx"
PLENO_PATH = "data/MATER_Pleno_1T2026_18.xlsx"

COSTA_CAP = 200  # Costa Atlántica comparte L6=200MW solar con PBA

GHI_CORREDOR = {
    "PBA CENTRO-SUR": 1570, "COSTA ATLÁNTICA": 1560,
    "LITORAL": 1500,        "MISIONES": 1390,
    "NEA": 1640,            "GBA": 1390,
    "BS. AS.": 1470,        "COMAHUE": 1650,
    "PATAGONIA": 1700,      "NOA": 1900,
    "CENTRO": 1700,         "CUYO": 2000,
}

GHI_NODO = {
    "1142":1550,"1260":1600,"1270":1600,"1051":1510,"1052":1510,"1053":1510,"1054":1510,
    "1184":1570,"1183":1570,"1187":1570,"1186":1570,"1188":1580,"1185":1580,"1189":1580,
    "1193":1530,"1197":1530,"1195":1530,"1191":1530,"1199":1520,"1141":1620,"1003":1620,
    "1137":1590,"1012":1600,"1013":1600,"1240":1600,"1243":1600,"1230":1610,"1233":1610,
    "1181":1580,"1182":1580,"1140":1560,"1163":1560,"1159":1560,"1244":1560,"1145":1570,
    "1146":1570,"1165":1570,"1171":1570,"1250":1600,"1160":1600,"1161":1590,"1174":1590,
    "1176":1610,"1147":1580,"1148":1580,"1150":1580,"1200":1590,"1212":1590,"1220":1600,
    "1173":1560,"1151":1560,"1135":1550,"1153":1550,"1154":1550,"1155":1550,"1158":1570,
    "1167":1570,"1169":1570,"1172":1570,"1178":1570,"1179":1570,"2047":1570,"1046":1580,
    "1047":1580,"1055":1580,"1056":1580,"1130":1580,"1132":1570,"1136":1550,"1133":1550,
    "1134":1550,"1156":1550,"1157":1550,"1166":1550,"9464":1510,"1138":1590,"1139":1620,
    "1143":1620,"1192":1530,"1194":1530,"1196":1530,"1198":1520,
    "2000":1590,"2006":1590,"2030":1590,"2004":1600,"2005":1600,"2032":1600,"2033":1600,
    "2035":1600,"2036":1600,"2039":1600,"2044":1560,"2045":1560,"2046":1560,"2010":1520,
    "2014":1520,"2015":1520,"2016":1520,"2017":1520,"2018":1520,"2019":1520,"2020":1590,
    "2038":1590,"2031":1530,"2037":1530,"2040":1530,"2041":1530,"2042":1530,"2043":1530,
    "2001":1510,"2002":1510,"2003":1510,"2009":1510,"2011":1510,"2012":1510,"2013":1510,
    "7048":1480,"7070":1480,"7077":1500,"7078":1500,"7079":1500,"7072":1490,"7081":1490,
    "7082":1490,"7083":1490,"7084":1490,"7085":1490,"7020":1500,"7034":1500,"7010":1510,
    "7019":1510,"7000":1520,"7016":1520,"7076":1520,"7014":1510,"7074":1510,"7075":1510,
    "7013":1520,"7060":1520,"7057":1490,"7059":1490,"7006":1490,"7029":1490,"7031":1490,
    "7032":1490,"7058":1480,"7003":1490,"7004":1490,"7053":1490,"7054":1490,"7021":1520,
    "7023":1520,"7025":1520,"7027":1520,"7050":1520,"7001":1520,
    "8001":1380,"8002":1380,"8006":1380,"8007":1380,"8027":1380,"8028":1380,"8004":1380,
    "8050":1380,"8046":1400,"8049":1400,"8024":1400,"8025":1400,"8014":1390,"8015":1390,
    "8011":1390,"8012":1390,"8017":1390,
    "8030":1600,"8089":1650,"8047":1610,"8060":1610,"8125":1610,"8126":1610,"8116":1620,
    "8117":1620,"8113":1620,"8114":1620,"8115":1620,"8085":1650,"8086":1650,"8087":1650,
    "8077":1640,"8078":1640,"8079":1640,"8081":1640,"8082":1640,"8083":1640,"8091":1680,
    "8092":1680,"8032":1680,"8095":1680,"8157":1680,"8158":1680,"8038":1680,"8039":1680,
    "8048":1680,"8100":1680,"8097":1680,"8098":1680,"8033":1680,"8020":1680,"8021":1680,
    "8103":1680,"8110":1680,"8111":1680,"8107":1680,"8108":1680,"8040":1620,"8139":1620,
    "8140":1620,"8133":1620,"8134":1620,"8135":1620,"8130":1620,"8131":1620,"8127":1620,
    "8128":1620,"8136":1620,"8137":1620,"8142":1620,"8152":1620,"8155":1620,"8042":1620,
    "8153":1620,"8045":1620,"8154":1620,"8143":1610,"8144":1610,"8149":1610,"8150":1610,
    "8146":1610,"8147":1610,"8062":1660,"8063":1660,"8056":1660,"8057":1660,"8059":1660,
    "8052":1660,"8061":1660,"8073":1670,"8074":1670,"8067":1670,"8072":1670,"8068":1670,
    "8069":1670,"8167":1640,"8168":1640,"8169":1640,"8000":1680,"8164":1680,"8036":1680,
    "8119":1680,"8122":1680,"8034":1680,"8160":1680,"8161":1680,
    "9010":1390,"9016":1390,"9020":1390,"9030":1390,"9017":1390,"9018":1390,"9022":1390,
    "9451":1420,"9524":1450,"9525":1450,"9530":1500,"9462":1430,"9463":1430,"9461":1430,
    "9000":1450,"9454":1450,"9455":1450,"9456":1470,"9457":1470,"9458":1480,"9459":1480,
    "9460":1490,"9520":1480,"9006":1460,"9007":1460,"9008":1460,"9009":1460,"7030":1480,
    "9005":1480,"9012":1490,"9013":1490,"9014":1510,"9015":1490,"9500":1480,"9511":1480,
    "9515":1480,"9517":1480,"9003":1470,"9004":1470,"9510":1480,"9516":1470,"9450":1420,
    "9453":1430,
    "1033":1650,"1034":1650,"1035":1650,"1042":1650,"1098":1650,"1104":1650,"1105":1650,
    "1106":1650,"1123":1650,"1050":1600,"1066":1650,"1124":1600,"1011":1600,"1020":1600,
    "4009":2100,"4013":2100,"4291":1900,"4300":1900,"4310":1900,"4280":1900,
    "4080":1850,"4038":1850,"4200":1850,"4125":1800,"4117":1800,
    "5061":1700,"5160":1700,"5140":1700,"5057":1750,
    "6240":1950,"6250":1950,"6400":2050,"6430":2050,"6431":2050,"6330":2050,"6181":2000,
}

COORDS = {
    "1142":(-36.9,-60.3),"1260":(-38.7,-62.3),"1184":(-36.8,-59.8),"1187":(-36.8,-59.3),
    "1189":(-37.6,-59.5),"1193":(-35.6,-59.8),"1141":(-38.0,-60.1),"1230":(-38.9,-62.1),
    "1250":(-38.7,-62.3),"1181":(-37.1,-61.0),"1163":(-38.1,-62.2),"1145":(-37.9,-63.0),
    "1165":(-37.8,-63.0),"1171":(-38.1,-62.2),"1012":(-38.5,-62.4),"1240":(-38.5,-62.3),
    "1243":(-38.5,-62.2),"1137":(-37.6,-60.1),"1046":(-38.2,-62.6),"1145":(-37.9,-63.0),
    "2030":(-37.5,-57.1),"2041":(-38.0,-57.5),"2020":(-38.6,-58.7),"2032":(-37.9,-58.2),
    "2044":(-37.3,-56.9),"2001":(-37.1,-56.9),"2010":(-38.3,-57.8),
    "8113":(-29.2,-58.8),"8030":(-27.5,-58.8),"8167":(-27.5,-59.0),"8089":(-26.2,-58.2),
    "8116":(-30.2,-59.3),"8047":(-28.5,-59.1),"8000":(-27.3,-60.5),"8161":(-27.3,-60.4),
    "8160":(-27.3,-60.4),"8073":(-26.0,-58.6),"8036":(-27.0,-60.2),"8034":(-26.8,-60.1),
    "8119":(-26.9,-60.2),"8122":(-26.9,-60.3),"8056":(-23.9,-61.8),"8059":(-24.1,-60.6),
    "8045":(-28.1,-56.1),
    "7081":(-31.4,-58.0),"7082":(-31.4,-58.0),"7058":(-31.6,-60.7),"7034":(-33.2,-61.3),
    "7077":(-29.9,-60.1),"7060":(-28.5,-59.7),"7010":(-33.8,-61.9),"7019":(-33.5,-61.5),
    "7000":(-34.5,-62.7),"7016":(-29.2,-59.6),"7076":(-29.1,-59.6),"7057":(-31.4,-60.9),
    "9016":(-35.0,-58.5),"9010":(-34.8,-58.6),"9020":(-34.9,-57.9),"9017":(-35.0,-57.9),
    "9006":(-33.5,-60.0),"9007":(-33.5,-60.0),"9008":(-34.0,-59.1),"9009":(-34.0,-59.1),
    "9015":(-35.1,-60.5),"9458":(-34.9,-60.0),"9460":(-35.4,-60.2),"9456":(-34.7,-59.4),
    "9012":(-34.6,-60.9),"9013":(-34.6,-60.9),"9014":(-34.8,-61.5),"9462":(-34.2,-58.9),
    "9000":(-34.6,-59.1),"9454":(-34.6,-59.1),"7030":(-34.1,-61.0),"9005":(-34.0,-61.1),
    "9500":(-33.9,-60.6),"9030":(-34.7,-58.3),
    "1033":(-35.8,-64.2),"1034":(-36.2,-64.3),"1035":(-35.7,-63.7),"1042":(-34.8,-64.4),
    "1098":(-37.5,-63.8),"1104":(-37.4,-65.0),"1105":(-38.3,-64.6),"1106":(-38.7,-67.3),
    "1123":(-37.4,-64.3),"1050":(-39.2,-70.2),"1066":(-38.7,-67.3),"1124":(-40.2,-70.7),
    "1011":(-38.9,-68.8),"1020":(-39.2,-70.2),
    "3000":(-43.3,-65.1),"3010":(-46.0,-67.5),"3020":(-51.6,-69.2),"3030":(-40.8,-62.9),
    "3040":(-53.8,-67.7),"3050":(-41.1,-71.3),"3060":(-42.8,-65.0),
    "4009":(-24.2,-65.3),"4013":(-24.2,-65.3),"4291":(-26.8,-65.2),"4300":(-27.0,-65.3),
    "4310":(-27.5,-65.0),"4280":(-27.5,-65.3),"4080":(-29.4,-66.3),"4038":(-29.3,-67.2),
    "4200":(-28.0,-63.9),"4125":(-27.7,-63.0),"4117":(-26.6,-62.8),
    "5061":(-33.7,-65.5),"5160":(-32.2,-63.6),"5140":(-33.5,-63.0),"5057":(-33.5,-66.4),
    "6240":(-34.6,-68.4),"6250":(-32.9,-68.8),"6400":(-31.5,-68.5),"6430":(-31.7,-68.2),
    "6431":(-31.4,-68.4),"6330":(-32.5,-68.0),"6181":(-33.5,-69.0),
}

COLOR_MAP = {
    "PBA CENTRO-SUR":"#2E75B6","COSTA ATLÁNTICA":"#1D9E75","LITORAL":"#3B6D11",
    "MISIONES":"#BA7517","NEA":"#D85A30","GBA":"#534AB7","BS. AS.":"#993556",
    "COMAHUE":"#1A7BAF","PATAGONIA":"#0D4C73","NOA":"#C84B11","CENTRO":"#5C8A1E","CUYO":"#8B1A8B",
}


# ── PARSER ────────────────────────────────────────────────────────────────────
def extract_mw(text):
    if not text or str(text).strip() in ["", "None"]:
        return None
    t = str(text).replace("\n", " ")
    m1 = re.search(r"\+\s*(\d+)\s*MW\s*\(#1\)", t)
    if m1:
        return int(m1.group(1))
    m = re.findall(r"(\d+)\s*MW", t)
    return int(m[0]) if m else None


@st.cache_data(show_spinner=False)
def parse_anexo3(refa_path, pleno_path):
    import openpyxl

    def is_export_limit(text):
        if not text or str(text).strip() in ["", "None"]: return False
        t = str(text).upper()
        # "EXPORTACI" cubre tanto EXPORTACION como EXPORTACIÓN (con acento)
        return "EXPORTACI" in t or "CORREDOR PATAGONIA" in t

    def get_corridor_caps(ws):
        """Primera pasada: techo de exportación solar de cada corredor.
        Si tiene límite de exportación explícito → ese es el techo absoluto (puede ser 0).
        Si no tiene → sin techo de corredor (limitado solo por PDI/límites propios)."""
        caps = {}
        corredor_actual = ""
        for row in list(ws.iter_rows(values_only=True))[2:]:
            c0 = str(row[0]).strip() if row[0] else ""
            if c0 and c0 != "None":
                corredor_actual = c0.upper().strip()
            if not row[1] or corredor_actual in caps:
                continue
            for col_idx in range(6, 12):
                raw = str(row[col_idx]) if row[col_idx] else ""
                if is_export_limit(raw):
                    caps[corredor_actual] = extract_mw(raw)
                    break
        return caps

    def leer(filepath):
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb["ANEXO 3.1"]
        corridor_caps = get_corridor_caps(ws)
        resultado = {}
        corredor_actual = ""
        last_lim = {i: None for i in range(1, 7)}
        for row in list(ws.iter_rows(values_only=True))[2:]:
            c0 = str(row[0]).strip() if row[0] else ""
            if c0 and c0 != "None":
                corredor_actual = c0.upper().strip()
                last_lim = {i: None for i in range(1, 7)}
            id_ = str(row[1]) if row[1] else ""
            if not id_ or id_ == "None":
                continue
            tipo = str(row[3]) if row[3] else ""
            for nivel in range(1, 7):
                val = extract_mw(str(row[5 + nivel]) if row[5 + nivel] else "")
                if val is not None:
                    last_lim[nivel] = val
            try:
                pdi = int(row[5]) if row[5] and str(row[5]) not in ["", "None", "-"] else None
            except (ValueError, TypeError):
                pdi = None
            lims = [x for x in [pdi] + [last_lim[i] for i in range(1, 7)] if x is not None]
            corr_cap = corridor_caps.get(corredor_actual)
            if corr_cap is not None:
                lims.append(corr_cap)
            if "COSTA" in corredor_actual:
                lims.append(COSTA_CAP)
            cap = min(lims) if lims else 0
            resultado[id_] = {"corredor": corredor_actual, "nombre": str(row[2]) if row[2] else "",
                               "tipo": tipo, "cap": cap}
        return resultado

    refa  = leer(refa_path)
    pleno = leer(pleno_path)
    rows = []
    for id_, d in refa.items():
        if d["tipo"] != "EETT":
            continue
        cap_r = d["cap"]
        cap_p = pleno.get(id_, {}).get("cap", 0)
        corr  = d["corredor"]
        ghi   = GHI_NODO.get(id_) or GHI_CORREDOR.get(corr, 1500)
        coords = COORDS.get(id_, (None, None))
        rows.append({
            "id": id_, "nombre": d["nombre"], "corredor": corr,
            "cap_refa": cap_r, "cap_pleno": cap_p, "ghi": ghi,
            "lat": coords[0], "lon": coords[1],
            "tiene_cap": cap_r > 0,
        })
    return pd.DataFrame(rows)


with st.spinner("Leyendo datos de Cammesa..."):
    try:
        df_raw = parse_anexo3(REFA_PATH, PLENO_PATH)
        load_error = None
    except Exception as e:
        df_raw = None
        load_error = str(e)

if load_error or df_raw is None:
    st.error(f"No se pudo leer el Excel: {load_error}")
    st.info("Asegurate de que los archivos estén en `data/` del repo con los nombres exactos.")
    st.stop()


# ── RANKING ───────────────────────────────────────────────────────────────────
def compute(df_base, w_ghi, w_mw, umbral):
    df = df_base.copy()
    activos = df[df["tiene_cap"]].copy()
    if activos.empty:
        df["score"] = None; df["rk"] = None; df["rk_ghi"] = None
        df["segmento"] = "📋 Sin capacidad hoy"
        return df
    gn_min, gn_max = activos["ghi"].min(), activos["ghi"].max()
    mw_min, mw_max = activos["cap_refa"].min(), activos["cap_refa"].max()
    def norm(s, mn, mx):
        return (s - mn) / (mx - mn) if mx != mn else 1.0
    activos["score"] = ((w_ghi/100) * norm(activos["ghi"], gn_min, gn_max)
                      + (w_mw /100) * norm(activos["cap_refa"], mw_min, mw_max))
    activos["rk"]     = activos["score"].rank(ascending=False, method="min").astype(int)
    activos["rk_ghi"] = activos["ghi"].rank(ascending=False, method="min").astype(int)
    activos["segmento"] = activos["cap_refa"].apply(
        lambda c: "⭐ Desarrollar (AURA)" if c <= umbral else "🤝 Ofrecer a terceros")
    sin_cap = df[~df["tiene_cap"]].copy()
    for col in ["score","rk","rk_ghi"]: sin_cap[col] = None
    sin_cap["segmento"] = "📋 Sin capacidad hoy"
    return pd.concat([activos, sin_cap], ignore_index=True)


# ── SIDEBAR ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## ⚡ AURA Energy")
    st.markdown("**MATER Dashboard — Argentina**")
    st.caption(f"{len(df_raw)} nodos EETT · 12 corredores")
    st.markdown("---")
    st.markdown("### Pesos del ranking")
    w_ghi = st.slider("Peso GHI (recurso solar)", 0, 100, 80, 5)
    w_mw  = 100 - w_ghi
    st.markdown(f"Peso MW disponibles: **{w_mw}%**")
    st.markdown("---")
    st.markdown("### Filtros")
    mostrar_sin_cap = st.toggle("Ver nodos sin capacidad hoy", value=False)
    todos_corr = ["Todos"] + sorted(df_raw["corredor"].unique().tolist())
    sel_corr   = st.selectbox("Corredor", todos_corr)
    rng_ghi    = st.slider("GHI mínimo", int(df_raw.ghi.min()), int(df_raw.ghi.max()),
                           int(df_raw.ghi.min()))
    st.markdown("---")
    st.markdown("### Segmentación AURA")
    umbral = st.slider("Umbral MW — Desarrollar vs Ofrecer", 10, 200, 50, 10)
    st.caption(f"≤{umbral} MW → AURA desarrolla  ·  >{umbral} MW → Ofrecer a terceros")

df_all = compute(df_raw, w_ghi, w_mw, umbral)

mask = df_all["ghi"] >= rng_ghi
if sel_corr != "Todos": mask &= (df_all["corredor"] == sel_corr)
if not mostrar_sin_cap: mask &= df_all["tiene_cap"]
df_f = df_all[mask].copy()

# ── HEADER + MÉTRICAS ─────────────────────────────────────────────────────────
st.markdown("# ⚡ AURA Energy — Oportunidades MATER")
df_act = df_f[df_f["tiene_cap"]]
total_mw = int(df_act["cap_refa"].sum())
top_n    = df_act.loc[df_act["rk"].idxmin(), "nombre"] if not df_act.empty else "—"
n_aura   = (df_f["segmento"] == "⭐ Desarrollar (AURA)").sum()
n_terc   = (df_f["segmento"] == "🤝 Ofrecer a terceros").sum()
st.markdown(f"**{len(df_f)} nodos visibles** ({len(df_act)} con capacidad) · GHI {w_ghi}% / MW {w_mw}%")

m1,m2,m3,m4 = st.columns(4)
with m1: st.markdown(f'<div class="metric-card"><div class="metric-lbl">Total MW Ref A</div><div class="metric-val">{total_mw:,}</div></div>', unsafe_allow_html=True)
with m2: st.markdown(f'<div class="metric-card"><div class="metric-lbl">Nodo #1 del ranking</div><div class="metric-val" style="font-size:1rem;line-height:1.3">{top_n}</div></div>', unsafe_allow_html=True)
with m3: st.markdown(f'<div class="metric-card"><div class="metric-lbl">⭐ Para desarrollar</div><div class="metric-val">{n_aura}</div></div>', unsafe_allow_html=True)
with m4: st.markdown(f'<div class="metric-card"><div class="metric-lbl">🤝 Para ofrecer</div><div class="metric-val">{n_terc}</div></div>', unsafe_allow_html=True)
st.markdown("---")

tab1, tab2, tab3, tab4 = st.tabs(["📊 GHI vs MW", "🗺️ Mapa", "🏆 Ranking", "📋 Tabla completa"])

# ── TAB 1 ─────────────────────────────────────────────────────────────────────
with tab1:
    df_sc = df_f[df_f["tiene_cap"]].copy()
    st.markdown("### GHI estimado vs Capacidad disponible (Ref A)")
    if df_sc.empty:
        st.info("Sin nodos con capacidad para el filtro seleccionado.")
    else:
        fig = px.scatter(
            df_sc, x="ghi", y="cap_refa", color="corredor", size="cap_refa", size_max=48,
            symbol="segmento",
            symbol_map={"⭐ Desarrollar (AURA)":"star","🤝 Ofrecer a terceros":"circle"},
            color_discrete_map=COLOR_MAP, hover_name="nombre",
            hover_data={"id":True,"corredor":True,"ghi":True,"cap_refa":True,
                        "score":":.3f","rk":True,"segmento":True},
            labels={"ghi":"GHI (kWh/m²/año)","cap_refa":"Cap. Ref A (MW)",
                    "corredor":"Corredor","rk":"Ranking"},
        )
        fig.add_hline(y=umbral, line_dash="dot", line_color="#BA7517",
                      annotation_text=f"Umbral {umbral} MW", annotation_position="right")
        fig.add_vline(x=1550, line_dash="dot", line_color="#aaa",
                      annotation_text="GHI 1550", annotation_position="top right")
        fig.update_layout(height=500, plot_bgcolor="white", paper_bgcolor="white",
                          font=dict(family="Arial"),
                          legend=dict(orientation="v", x=1.01, y=1),
                          xaxis=dict(showgrid=True, gridcolor="#F0F0F0"),
                          yaxis=dict(showgrid=True, gridcolor="#F0F0F0"))
        st.plotly_chart(fig, use_container_width=True)

    by_c = df_act.groupby("corredor")["cap_refa"].sum().reset_index().sort_values("cap_refa", ascending=True)
    fig_b = px.bar(by_c, x="cap_refa", y="corredor", orientation="h",
                   color="corredor", color_discrete_map=COLOR_MAP,
                   labels={"cap_refa":"MW totales","corredor":""}, text="cap_refa")
    fig_b.update_traces(texttemplate="%{text} MW", textposition="outside")
    fig_b.update_layout(height=340, showlegend=False, plot_bgcolor="white", paper_bgcolor="white",
                        xaxis=dict(showgrid=True, gridcolor="#F0F0F0"))
    st.plotly_chart(fig_b, use_container_width=True)

# ── TAB 2 ─────────────────────────────────────────────────────────────────────
with tab2:
    df_map = df_f.dropna(subset=["lat","lon"]).copy()
    df_map["tamaño_mapa"] = df_map["cap_refa"].clip(lower=1)
    st.markdown(f"### Mapa — {len(df_map)} nodos con coordenadas")
    if df_map.empty:
        st.info("Sin coordenadas para el filtro seleccionado.")
    else:
        fig_m = px.scatter_mapbox(
            df_map, lat="lat", lon="lon", color="corredor", size="tamaño_mapa", size_max=28,
            color_discrete_map=COLOR_MAP, hover_name="nombre",
            hover_data={"id":True,"corredor":True,"ghi":True,"cap_refa":True,
                        "segmento":True,"lat":False,"lon":False,"tamaño_mapa":False},
            zoom=3.8, center={"lat":-35,"lon":-64},
            mapbox_style="open-street-map", opacity=0.85,
            labels={"cap_refa":"MW Ref A","corredor":"Corredor"},
        )
        fig_m.update_layout(height=580, margin={"r":0,"t":0,"l":0,"b":0},
                             legend=dict(orientation="v",x=0.01,y=0.99,
                                         bgcolor="rgba(255,255,255,0.85)"))
        st.plotly_chart(fig_m, use_container_width=True)

# ── TAB 3 ─────────────────────────────────────────────────────────────────────
with tab3:
    col_a, col_b = st.columns(2)
    with col_a:
        st.markdown("### ⭐ Para desarrollar directamente (AURA)")
        st.caption(f"≤ {umbral} MW · top ranking ponderado")
        df_aura = df_f[df_f["segmento"]=="⭐ Desarrollar (AURA)"].sort_values("rk").head(30)
        if df_aura.empty:
            st.info("Sin nodos en este segmento.")
        for _, r in df_aura.iterrows():
            cc = COLOR_MAP.get(r["corredor"],"#888")
            st.markdown(f"""<div style="border-left:3px solid {cc};padding:6px 10px;margin:3px 0;
                background:#FAFAFA;border-radius:0 6px 6px 0">
                <span style="font-weight:600;color:#1F3864">#{int(r['rk'])} {r['nombre']}</span>
                <span style="font-size:.77rem;color:#888;margin-left:6px">{r['corredor']}</span><br>
                <span style="font-size:.82rem">GHI <b>{r['ghi']}</b> · <b>{int(r['cap_refa'])} MW</b>
                · score {r['score']:.3f}</span></div>""", unsafe_allow_html=True)
    with col_b:
        st.markdown("### 🤝 Para ofrecer a terceros como consultores")
        st.caption(f"> {umbral} MW · buen GHI · top ranking ponderado")
        df_terc = df_f[df_f["segmento"]=="🤝 Ofrecer a terceros"].sort_values("rk").head(30)
        if df_terc.empty:
            st.info("Sin nodos en este segmento.")
        for _, r in df_terc.iterrows():
            cc = COLOR_MAP.get(r["corredor"],"#888")
            st.markdown(f"""<div style="border-left:3px solid {cc};padding:6px 10px;margin:3px 0;
                background:#F0FAF5;border-radius:0 6px 6px 0">
                <span style="font-weight:600;color:#0F6E56">#{int(r['rk'])} {r['nombre']}</span>
                <span style="font-size:.77rem;color:#888;margin-left:6px">{r['corredor']}</span><br>
                <span style="font-size:.82rem">GHI <b>{r['ghi']}</b> · <b>{int(r['cap_refa'])} MW</b>
                · score {r['score']:.3f}</span></div>""", unsafe_allow_html=True)
    if mostrar_sin_cap:
        st.markdown("---")
        st.markdown("### 📋 Sin capacidad hoy — pueden abrirse en futuros trimestres")
        df_sin = df_f[df_f["segmento"]=="📋 Sin capacidad hoy"].sort_values(
            ["corredor","ghi"], ascending=[True,False]).head(50)
        for _, r in df_sin.iterrows():
            cc = COLOR_MAP.get(r["corredor"],"#888")
            st.markdown(f"""<div style="border-left:3px solid {cc};padding:4px 10px;margin:2px 0;
                background:#F8F8F8;border-radius:0 6px 6px 0;opacity:0.65">
                <span style="color:#555">{r['nombre']}</span>
                <span style="font-size:.75rem;color:#aaa;margin-left:6px">{r['corredor']}</span>
                <span style="font-size:.75rem;color:#888;float:right">GHI {r['ghi']}</span>
                </div>""", unsafe_allow_html=True)

# ── TAB 4 ─────────────────────────────────────────────────────────────────────
with tab4:
    st.markdown("### Tabla completa")
    cols = ["rk","nombre","corredor","ghi","cap_refa","cap_pleno","score","segmento","id"]
    df_show = df_f[cols].copy()
    df_show.columns = ["Ranking","Nombre","Corredor","GHI","Cap. Ref A (MW)",
                        "Cap. Pleno (MW)","Score","Segmento","ID"]
    st.dataframe(
        df_show.sort_values("Ranking", na_position="last"),
        use_container_width=True, height=520,
        column_config={
            "Ranking": st.column_config.NumberColumn(format="%d"),
            "GHI": st.column_config.NumberColumn(format="%d kWh/m²"),
            "Cap. Ref A (MW)": st.column_config.NumberColumn(format="%d MW"),
            "Cap. Pleno (MW)": st.column_config.NumberColumn(format="%d MW"),
            "Score": st.column_config.NumberColumn(format="%.3f"),
        },
        hide_index=True,
    )

st.markdown("---")
st.caption("AURA Energy · Cammesa Anexo 3 Ref A 1T2026 · GHI: Global Solar Atlas · "
           "Para actualizar: reemplazá los archivos en la carpeta data/ del repo de GitHub")
