# -*- coding: utf-8 -*-
"""Exportable Excel del modelo financiero, un archivo por proyecto: hoja 'Supuestos'
(inputs editables) + hoja 'Flujo de Fondos' (formulas ano a ano, VAN/TIR/LCOE). Sin diseno
grafico - pensado para que el usuario lo baje y lo siga afinando el mismo en Excel.

Estilo y formulas replican el formato del modelo de referencia que uso AURA (Flujo_Fondos_
Solar_20MW): inputs en azul con fondo amarillo, formulas en negro, todo calculado por formula
(no valores fijos) para que recalcule solo si se cambia un supuesto. Extiende ese modelo con las
dos variables que el motor de aura.py sí modela y el de referencia no: el FC de recurso vs. FC
financiero (curtailment Ref A, 8%) y el recupero de credito fiscal de IVA en hasta 3 anios
(el de referencia asume IVA neutro en caja).
"""
import io
import zipfile

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

FONT_NAME = "Arial"
AZUL_INPUT = Font(name=FONT_NAME, size=10, color="0000FF")
NEGRO = Font(name=FONT_NAME, size=10, color="000000")
NEGRO_BOLD = Font(name=FONT_NAME, size=10, color="000000", bold=True)
LABEL_BOLD = Font(name=FONT_NAME, size=10, color="000000", bold=True)
TITULO = Font(name=FONT_NAME, size=14, color="000000", bold=True)
HEADER_ANIO = Font(name=FONT_NAME, size=11, color="FFFFFF", bold=True)
FILL_INPUT = PatternFill("solid", fgColor="FFFF00")
FILL_HEADER = PatternFill("solid", fgColor="1F4E5F")

FMT_PCT = "0.0%"
FMT_PCT2 = "0.00%"
FMT_MONEY = r'\$#,##0;"($"#,##0\);\-'
FMT_MONEY2 = r'\$#,##0.00'
FMT_NUM = "#,##0"


def _input_cell(ws, coord, value, fmt=None):
    c = ws[coord]
    c.value = value
    c.font = AZUL_INPUT
    c.fill = FILL_INPUT
    if fmt:
        c.number_format = fmt
    return c


def _formula_cell(ws, coord, formula, fmt=None, bold=False):
    c = ws[coord]
    c.value = formula
    c.font = NEGRO_BOLD if bold else NEGRO
    if fmt:
        c.number_format = fmt
    return c


def _label(ws, coord, texto, bold=True):
    c = ws[coord]
    c.value = texto
    if texto and texto[0] in ("=", "+", "-", "@"):
        c.data_type = "s"
    c.font = LABEL_BOLD if bold else NEGRO
    return c


def _nota(ws, coord, texto):
    c = ws[coord]
    c.value = texto
    if texto and texto[0] in ("=", "+", "-", "@"):
        # openpyxl interpreta cualquier string que arranca con estos caracteres como formula;
        # si el texto no es una formula valida, Excel se niega a abrir el archivo. Se fuerza
        # el tipo string explicitamente para evitarlo.
        c.data_type = "s"
    c.font = Font(name=FONT_NAME, size=9, italic=True, color="666666")
    return c


def _hoja_supuestos(wb, p, fs):
    ws = wb.create_sheet("Supuestos")
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 52

    ws["A1"] = f"{p['nombre']} ({'Solar' if p['tech'] == 'solar' else 'Eólico'}) - Supuestos del Modelo"
    ws["A1"].font = TITULO

    cap = p["cap"]
    fc_bruto = p.get("fc_bruto") if p.get("fc_bruto") is not None else p.get("fc_fin")
    fc_fin = p.get("fc_fin")
    curtailment_pct = 0.0
    if fc_bruto and fc_fin is not None and fc_bruto > 0:
        curtailment_pct = max(0.0, 1 - (fc_fin / fc_bruto))
    considerar_iva = fs.get("considerar_iva", True)
    tasa_iva = fs.get("tasa_iva", 0.21) if considerar_iva else 0.0

    _label(ws, "A3", "Capacidad (MW)")
    _input_cell(ws, "B3", cap)
    _label(ws, "A4", "Factor de planta (recurso)")
    _input_cell(ws, "B4", fc_bruto, FMT_PCT2)
    _nota(ws, "C4", "FC de recurso puro (sin descuento por curtailment Ref A)")
    _label(ws, "A5", "Curtailment Ref A")
    _input_cell(ws, "B5", curtailment_pct, FMT_PCT)
    _nota(ws, "C5", "% de energía no garantizada por despacho Ref A (0% si el nodo no depende de Ref A)")
    _label(ws, "A6", "Factor de planta financiero")
    _formula_cell(ws, "B6", "=B4*(1-B5)", FMT_PCT2)
    _nota(ws, "C6", "Fórmula: recurso x (1 - curtailment) - el que se usa en la generación de abajo")
    _label(ws, "A7", "Horas por año")
    _input_cell(ws, "B7", 8760)
    _label(ws, "A8", "Precio de energía (USD/MWh)")
    _input_cell(ws, "B8", fs["precio_mwh"])
    _label(ws, "A9", "CAPEX (USD/MW)")
    _input_cell(ws, "B9", fs["capex_mw"], FMT_MONEY)
    _label(ws, "A10", "CAPEX total (USD)")
    _formula_cell(ws, "B10", "=B3*B9", FMT_MONEY)
    _nota(ws, "C10", "Se desembolsa 100% en el Año 0")
    _label(ws, "A11", "OPEX (USD/MW/año)")
    _input_cell(ws, "B11", fs["opex_mw"], FMT_MONEY)
    _label(ws, "A12", "OPEX anual (USD)")
    _formula_cell(ws, "B12", "=B3*B11", FMT_MONEY)
    _nota(ws, "C12", "Fijo, sin escalación")
    _label(ws, "A13", "Tasa de descuento")
    _input_cell(ws, "B13", fs["tasa"], FMT_PCT)
    _label(ws, "A14", "Plazo del proyecto (años)")
    _input_cell(ws, "B14", fs["plazo"])
    _label(ws, "A15", "Degradación Año 1→2")
    _input_cell(ws, "B15", 0.01, FMT_PCT)
    _nota(ws, "C15", "Primer año (LID)")
    _label(ws, "A16", "Degradación años siguientes")
    _input_cell(ws, "B16", 0.004, FMT_PCT)
    _nota(ws, "C16", "Del año 3 en adelante")
    _label(ws, "A17", "Impuesto a las Ganancias")
    _input_cell(ws, "B17", 0.35, FMT_PCT)
    _label(ws, "A18", "Años de amortización (CAPEX)")
    _input_cell(ws, "B18", fs["amortizacion"])
    _nota(ws, "C18", "Lineal")
    _label(ws, "A19", "IVA (crédito fiscal)")
    _input_cell(ws, "B19", tasa_iva, FMT_PCT)
    _nota(ws, "C19", ("Crédito 100% recuperable: se compensa contra el débito fiscal de los años 1 y 2, "
                       "y si queda saldo se fuerza su recupero total en el año 3" if considerar_iva
                       else "No considerado en este escenario"))
    _label(ws, "A20", "Financiamiento")
    ws["B20"] = "100% Equity"
    ws["B20"].font = AZUL_INPUT
    ws["B20"].fill = FILL_INPUT
    _nota(ws, "C20", "Sin deuda")
    _label(ws, "A21", "Valor residual / desmantelamiento")
    _input_cell(ws, "B21", 0, FMT_MONEY)
    _nota(ws, "C21", "No aplica")
    return ws


def _hoja_flujo(wb, p, fs):
    ws = wb.create_sheet("Flujo de Fondos")
    plazo = int(fs["plazo"])
    ws["A1"] = f"Flujo de Fondos - {p['nombre']}"
    ws["A1"].font = TITULO

    ws.column_dimensions["A"].width = 30
    for i in range(plazo + 1):
        ws.column_dimensions[get_column_letter(2 + i)].width = 12

    filas = [
        ("Año", 3), ("Generación (MWh)", 4), ("Ingresos (USD)", 5), ("OPEX (USD)", 6),
        ("Amortización (USD)", 7), ("EBT (USD)", 8), ("Impuesto a las Ganancias (USD)", 9),
        ("Utilidad Neta (USD)", 10), ("CAPEX (USD)", 11), ("Desembolso crédito IVA (USD)", 12),
        ("Crédito IVA remanente (USD)", 13), ("Recupero crédito IVA (USD)", 14),
        ("Flujo de Fondos (USD)", 15),
    ]
    for label, row in filas:
        _label(ws, f"A{row}", label, bold=(row in (4, 15)))

    col_b = "B"
    col_last = get_column_letter(2 + plazo)

    # Año 0
    _formula_cell(ws, "B3", 0)
    ws["B3"].font = HEADER_ANIO
    ws["B3"].fill = FILL_HEADER
    ws["B3"].alignment = Alignment(horizontal="center")
    _formula_cell(ws, "B11", "=-Supuestos!$B$10", FMT_MONEY)
    _formula_cell(ws, "B12", "=-Supuestos!$B$10*Supuestos!$B$19", FMT_MONEY)
    _formula_cell(ws, "B13", "=-B12", FMT_MONEY)
    _formula_cell(ws, "B14", 0, FMT_MONEY)
    _formula_cell(ws, "B15", "=B11+B12", FMT_MONEY, bold=True)

    for i in range(1, plazo + 1):
        col = get_column_letter(2 + i)
        prev = get_column_letter(1 + i)
        _formula_cell(ws, f"{col}3", i)
        ws[f"{col}3"].font = HEADER_ANIO
        ws[f"{col}3"].fill = FILL_HEADER
        ws[f"{col}3"].alignment = Alignment(horizontal="center")

        if i == 1:
            gen_formula = "=Supuestos!$B$3*Supuestos!$B$7*Supuestos!$B$6"
        elif i == 2:
            gen_formula = f"={prev}4*(1-Supuestos!$B$15)"
        else:
            gen_formula = f"={prev}4*(1-Supuestos!$B$16)"
        _formula_cell(ws, f"{col}4", gen_formula, FMT_NUM)
        _formula_cell(ws, f"{col}5", f"={col}4*Supuestos!$B$8", FMT_MONEY)
        _formula_cell(ws, f"{col}6", "=-Supuestos!$B$12", FMT_MONEY)
        _formula_cell(ws, f"{col}7", f'=IF({col}$3<=Supuestos!$B$18,-Supuestos!$B$10/Supuestos!$B$18,0)', FMT_MONEY)
        _formula_cell(ws, f"{col}8", f"={col}5+{col}6+{col}7", FMT_MONEY)
        _formula_cell(ws, f"{col}9", f"=-MAX({col}8,0)*Supuestos!$B$17", FMT_MONEY)
        _formula_cell(ws, f"{col}10", f"={col}8+{col}9", FMT_MONEY)
        _formula_cell(ws, f"{col}13", f"={prev}13-{prev}14", FMT_MONEY)
        _formula_cell(ws, f"{col}14",
                       f'=IF({col}13<=0,0,IF({col}$3<3,MIN({col}5*Supuestos!$B$19,{col}13),{col}13))', FMT_MONEY)
        _formula_cell(ws, f"{col}15", f"={col}10-{col}7+{col}11+{col}12+{col}14", FMT_MONEY, bold=True)

    _label(ws, "A17", "VAN (@ tasa de descuento)")
    _formula_cell(ws, "B17", f"=NPV(Supuestos!$B$13,C15:{col_last}15)+B15", FMT_MONEY, bold=True)
    ws["B17"].fill = FILL_INPUT
    _label(ws, "A18", "TIR")
    _formula_cell(ws, "B18", f"=IRR(B15:{col_last}15)", FMT_PCT2, bold=True)
    ws["B18"].fill = FILL_INPUT
    _label(ws, "A19", "LCOE (USD/MWh)")
    _formula_cell(ws, "B19",
                  f"=(-B11+NPV(Supuestos!$B$13,C6:{col_last}6)*-1)/NPV(Supuestos!$B$13,C4:{col_last}4)",
                  FMT_MONEY2, bold=True)
    ws["B19"].fill = FILL_INPUT
    _nota(ws, "C19", "LCOE = (CAPEX + VAN de OPEX) / VAN de la generación - costo puro del activo, sin IVA ni efecto fiscal")

    ws.freeze_panes = "B4"
    return ws


def generar_excel_modelo(p):
    """Genera el .xlsx (bytes) del modelo financiero de un proyecto: hoja Supuestos + hoja
    Flujo de Fondos, todo por formula. `p` es el dict de proyecto (igual que el que usa el PDF),
    necesita fin_snapshot (supuestos financieros con los que se calculo)."""
    fs = p.get("fin_snapshot")
    if not fs:
        raise ValueError("El proyecto no tiene fin_snapshot - no se puede armar el modelo.")
    wb = Workbook()
    wb.remove(wb.active)
    _hoja_supuestos(wb, p, fs)
    _hoja_flujo(wb, p, fs)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _nombre_archivo(p, sufijo=".xlsx"):
    import re
    base = re.sub(r'[\\/:*?"<>|]', "", p["nombre"]).strip()
    return f"modelo_financiero_{base}{sufijo}"


def generar_excel_o_zip(proyectos):
    """Si hay 1 proyecto devuelve (nombre.xlsx, bytes) directo. Si hay mas de uno, devuelve
    (nombre.zip, bytes) con un .xlsx por proyecto adentro."""
    validos = [p for p in proyectos if p.get("fin_snapshot")]
    if not validos:
        return None, None
    if len(validos) == 1:
        p = validos[0]
        return _nombre_archivo(p), generar_excel_modelo(p)
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for p in validos:
            zf.writestr(_nombre_archivo(p), generar_excel_modelo(p))
    return "modelos_financieros_aura.zip", buf.getvalue()
